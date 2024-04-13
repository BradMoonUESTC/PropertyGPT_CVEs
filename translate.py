import openpyxl
import openai
import requests
# 设置OpenAI API密钥
openai.api_key = ''
from tqdm import tqdm
def translate_text(prompt):
        api_key = ''  # Replace with your actual OpenAI API key
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}"
        }
        data = {
            # "model": "gpt-4-turbo-preview",
            "model":"gpt-3.5-turbo-0125",
            "messages": [
                {
                    "role": "user",
                    "content": prompt
                }
            ]
        }
        response = requests.post('https://api.openai.com/v1/chat/completions', headers=headers, json=data)
        response_josn = response.json()
        if 'choices' not in response_josn:
            return ''
        return response_josn['choices'][0]['message']['content']

def translate_excel(file_path):
    """
    读取Excel文件并翻译所有单元格中的文本
    """
    wb = openpyxl.load_workbook(file_path)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in tqdm(sheet.iter_rows()):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    translated_text = translate_text("翻译以下内容为英文:"+cell.value)
                    cell.value = translated_text
    wb.save(file_path.replace('.xlsx', '_translated.xlsx'))

# Excel文件路径
file_path = '工作簿6.xlsx'
translate_excel(file_path)
